"""Log Worksheet Structure
SessionId	ConversationId	UtteranceIndex	Utterance	Role	Type	CreatedBy	CreationTime
"""
import pandas as pd
from graphapi import construct_range_for_update
from openpyxl.utils import get_column_letter


def create_conversation_log(messages):
    """
    Creates a conversation log DataFrame from a messages dictionary.

    Args:
        messages (dict): A dictionary containing the conversation messages.

    Returns:
        pandas.DataFrame: A DataFrame containing the conversation log records.

    Example:
        messages = {
            "cid": "conversation_id",
            "user": "username",
            "SessionId": "session_id",
            "timestamp": "2023-05-20T10:30:00",
            "useForTraining": "yes",
            "data": [
                {
                    "message": "Hello!",
                    "role": "user"
                },
                {
                    "message": "Hi there!",
                    "role": "assistant"
                }
            ]
        }
        log_df = create_conversation_log(messages)
    """
    records = []
    cid = messages["cid"]
    username = messages["user"]
    sid = messages["SessionId"]
    timestamp = messages["timestamp"]
    use_for_training = messages.get("useForTraining", "NA")
    for i, msg in enumerate(messages["data"]):
        record = {
            "SessionId": sid,
            "ConversationId": cid,
            "UtteranceIndex": i,
            "Utterance": msg["message"].replace("|", "\n"),
            "Role": msg["role"],
            "Type": "NA",
            "CreatedBy": username,
            "CreationTime": timestamp,
            "UseForTraining": use_for_training
        }
        if i == 0:
            _type = "system"
        else:
            _type = "query"
        text = msg["message"].lower()
        if "call" in text and "(" in text and ")" in text:
            _type = "api_call"

        record["Type"] = _type
        records.append(record)

    return pd.DataFrame(records)


def prepare_export_payload(data, nrows):
    """
    Prepares the export payload for conversation log data.

    This function takes a data dictionary and the desired number of rows to export as inputs.
    It prepares the export payload by creating a conversation log DataFrame from the data dictionary
    and constructing the range for updating the exported data.

    Args:
        data (dict): A dictionary containing the conversation data.
        nrows (int): The number of rows to export.

    Returns:
        tuple: A tuple containing the export payload, which includes the conversation log data as a list
               and the range for updating the exported data in the destination.

    Example:
        data = {
            "cid": "conversation_id",
            "user": "username",
            "SessionId": "session_id",
            "timestamp": "2023-05-20T10:30:00",
            "useForTraining": "yes",
            "data": [
                {
                    "message": "Hello!",
                    "role": "user"
                },
                {
                    "message": "Hi there!",
                    "role": "assistant"
                }
            ]
        }
        nrows = 10
        export_payload = prepare_export_payload(data, nrows)
    """
    df = create_conversation_log(data)
    alpha_range = construct_range_for_update(df, nrows)
    return df.values.tolist(), alpha_range


def map_indices_to_range_address(df, start_row=None, end_row=None):
    """
    Maps DataFrame indices to an Excel sheet range address.

    This function takes a DataFrame and optional start_row and end_row values as inputs.
    It maps the DataFrame indices to an Excel sheet range address by converting the indices
    and column letters to the appropriate format.

    Args:
        df (pandas.DataFrame): The DataFrame to map indices from.
        start_row (int): The starting row index. Default is None, which uses the first index in the DataFrame.
        end_row (int): The ending row index. Default is None, which uses the last index in the DataFrame.

    Returns:
        str: The Excel sheet range address corresponding to the DataFrame indices.

    Example:
        df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
        start_row = 2
        end_row = 5
        range_address = map_indices_to_range_address(df, start_row, end_row)
    """
    if start_row is None:
        start_row = df.index[0]
    if end_row is None:
        end_row = df.index[-1] + 2
    # Add 2 to account for header row and 0-based indexing
    start_row += 1
    end_row += 1
    # Add 1 to account for 0-based indexing
    start_column = get_column_letter(1)
    # Add 1 to account for 0-based indexing
    end_column = get_column_letter(len(df.columns))

    range_address = f"{start_column}{start_row}:{end_column}{end_row}"
    return range_address


def add_feedback_to_conversation(rows, data, nrows):
    """
    Adds feedback to the conversation log DataFrame.

    This function takes a list of rows, a data dictionary, and the desired number of rows to export as inputs.
    It creates a conversation log DataFrame from the list of rows, identifies the rows corresponding to the given
    conversation ID (cid), and updates the 'UseForTraining' column with the provided value from the data dictionary.
    Finally, it returns the updated portion of the DataFrame as a list and the range for updating the exported data.

    Args:
        rows (list): A list of rows representing the conversation log data.
        data (dict): A dictionary containing the conversation data.
        nrows (int): The number of rows to export.

    Returns:
        tuple: A tuple containing the updated portion of the conversation log data as a list
               and the range for updating the exported data in the destination.

    Example:
        rows = [
            ['SessionId', 'ConversationId', 'UtteranceIndex', 'Utterance', 'Role', 'Type', 'CreatedBy', 'CreationTime', 'UseForTraining'],
            ['session_id', 'conversation_id', 0, 'Hello!', 'user', 'NA', 'username', '2023-05-20T10:30:00', 'NA'],
            ['session_id', 'conversation_id', 1, 'Hi there!', 'assistant', 'NA', 'username', '2023-05-20T10:31:00', 'NA']
        ]
        data = {
            "cid": "conversation_id",
            "useForTraining": "yes"
        }
        nrows = 10
        updated_data, alpha_range = add_feedback_to_conversation(rows, data, nrows)
    """
    df = pd.DataFrame(rows[1:], columns=rows[0])
    cid = data["cid"]
    use_for_training = data["useForTraining"]
    indices = df[df.ConversationId == cid].index
    if len(indices) == 0:
        return prepare_export_payload(data, nrows)
    # set `UseForTraining` flag
    df.loc[indices, "UseForTraining"] = use_for_training
    # get update range
    alpha_range = map_indices_to_range_address(
        df, start_row=indices[0], end_row=indices[-1])
    return df.loc[indices].values.tolist(), alpha_range


def get_rows_by_cid(rows, conversation_id):
    """
    Retrieves rows from a conversation log DataFrame based on the conversation ID.

    This function takes a list of rows, representing the conversation log data, and a conversation ID as inputs.
    It creates a conversation log DataFrame from the list of rows and iterates over the rows to filter and extract
    the messages belonging to the given conversation ID. Each message is represented as a dictionary with 'role' and
    'textContent' attributes. The 'textContent' value is sanitized by replacing double quotes (") with single quotes (')
    to avoid any potential issues. Finally, it returns a list of messages extracted from the DataFrame.

    Args:
        rows (list): A list of rows representing the conversation log data.
        conversation_id (str): The conversation ID to filter the rows.

    Returns:
        list: A list of dictionaries representing the messages in the conversation, with 'role' and 'textContent' attributes.

    Example:
        rows = [
            ['SessionId', 'ConversationId', 'UtteranceIndex', 'Utterance', 'Role', 'Type', 'CreatedBy', 'CreationTime', 'UseForTraining'],
            ['session_id', 'conversation_id', 0, 'Hello!', 'user', 'NA', 'username', '2023-05-20T10:30:00', 'NA'],
            ['session_id', 'conversation_id', 1, 'Hi there!', 'assistant', 'NA', 'username', '2023-05-20T10:31:00', 'NA']
        ]
        conversation_id = 'conversation_id'
        messages = get_rows_by_cid(rows, conversation_id)
    """
    df = pd.DataFrame(rows[1:], columns=rows[0])
    messages = []
    for _, row in df[df.ConversationId == conversation_id].iterrows():
        messages.append({
            "role": row["Role"],
            "textContent": row["Utterance"].replace("\"", "'")
        })
    return messages
